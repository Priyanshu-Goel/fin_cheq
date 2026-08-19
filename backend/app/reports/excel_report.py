"""
Builds the downloadable Excel workbook: Summary, 5-Year Ratios, CAPM & Risk,
Red Flags, Backtest, Profit & Loss, Balance Sheet, Cash Flow, All Ratios
(Raw), and Raw Price Data sheets - the kind of structure a banking analyst
would expect to open and immediately navigate.

The Profit & Loss / Balance Sheet / Cash Flow / All Ratios sheets are the
*raw* tables fetched from indianapi.in/Screener ({row_label: {period:
value}}), not just the handful of curated metrics analysis/ratios.py
derives from them - the "5-Year Ratios" sheet only ever shows the named
metrics ratios.py knows to look for (ROE %, Debt to Equity, etc.); the raw
tables carry everything else that was actually fetched (Sales, Net Profit,
Debtor Days, Cash from Operating Activity, and so on) but had nowhere to
go before this.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

from app.models import AnalyzeResponse

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def _style_header_row(ws, row_idx: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)


def _add_raw_table_sheet(wb: Workbook, title: str, table: dict) -> None:
    """
    Renders a raw {row_label: {period: value}} table (the shape both
    fundamentals_api.py and screener_scraper.py produce) as its own sheet:
    one row per label, one column per period, sorted chronologically.
    Adds a near-empty placeholder sheet rather than skipping entirely when
    `table` is empty, so it's visibly "no data" rather than looking like
    the sheet was forgotten.
    """
    ws = wb.create_sheet(title)
    if not table:
        ws["A1"] = "No data available for this table."
        return

    all_periods = sorted({period for row in table.values() for period in row})
    ws.append(["Metric", *all_periods])
    _style_header_row(ws, 1, len(all_periods) + 1)
    for label, row in table.items():
        ws.append([label, *[row.get(p, "") for p in all_periods]])
    _autofit(ws, len(all_periods) + 1)


def build_excel_report(
    analysis: AnalyzeResponse, price_df: pd.DataFrame, fundamentals: dict, output_path: str
) -> str:
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{analysis.company_name} ({analysis.symbol}) - Equity Research Summary"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Executive Summary"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = analysis.summary
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A4:H12")
    ws.row_dimensions[4].height = 200

    ws["A14"] = "Risk Grade"
    ws["B14"] = analysis.risk.risk_grade
    ws["A15"] = "CAPM Cost of Equity"
    ws["B15"] = f"{analysis.capm.cost_of_equity:.2%}"
    ws["A16"] = "Beta"
    ws["B16"] = analysis.capm.beta
    _autofit(ws, 8)

    # --- Ratios sheet (curated: only the named metrics ratios.py derives -
    # see "All Ratios (Raw)" below for everything else that was fetched) ---
    ws2 = wb.create_sheet("Key Ratios (Curated)")
    all_years = sorted({yr for r in analysis.ratios for yr in r.values_by_year})
    ws2.append(["Metric", *all_years, "Trend"])
    _style_header_row(ws2, 1, len(all_years) + 2)
    for r in analysis.ratios:
        row = [r.metric] + [r.values_by_year.get(y, "") for y in all_years] + [r.trend]
        ws2.append(row)
    _autofit(ws2, len(all_years) + 2)

    # --- CAPM & Risk sheet ---
    ws3 = wb.create_sheet("CAPM & Risk")
    ws3.append(["Metric", "Value"])
    _style_header_row(ws3, 1, 2)
    capm_rows = [
        ("Beta", analysis.capm.beta),
        ("Risk-Free Rate", f"{analysis.capm.risk_free_rate:.2%}"),
        ("Market Return Assumption", f"{analysis.capm.market_return_assumption:.2%}"),
        ("Cost of Equity (CAPM)", f"{analysis.capm.cost_of_equity:.2%}"),
        ("Regression R-squared", analysis.capm.r_squared),
        ("Regression Window (yrs)", analysis.capm.regression_window_years),
        ("Annualized Volatility", f"{analysis.risk.annualized_volatility:.2%}"),
        ("Sharpe Ratio", analysis.risk.sharpe_ratio),
        ("Max Drawdown", f"{analysis.risk.max_drawdown:.2%}"),
        ("Value at Risk (95%, 1-day)", f"{analysis.risk.value_at_risk_95:.2%}"),
        ("Risk Grade", analysis.risk.risk_grade),
    ]
    for label, val in capm_rows:
        ws3.append([label, val])
    _autofit(ws3, 2)

    # --- Red Flags sheet ---
    ws4 = wb.create_sheet("Red Flags")
    ws4.append(["Severity", "Title", "Detail"])
    _style_header_row(ws4, 1, 3)
    for flag in analysis.red_flags:
        ws4.append([flag.severity, flag.title, flag.detail])
    _autofit(ws4, 3)

    # --- Backtest sheet ---
    ws5 = wb.create_sheet("Backtest")
    ws5.append(["Metric", "Value"])
    _style_header_row(ws5, 1, 2)
    bt = analysis.backtest
    for label, val in [
        ("As-of Date (signal generated)", bt.as_of_date),
        ("Predicted Expected Return (CAPM)", f"{bt.predicted_expected_return:.2%}"),
        ("Actual Realized Return (next 1yr)", f"{bt.actual_realized_return:.2%}"),
        ("Error", f"{bt.error:.2%}"),
        ("Directional Hit", bt.directional_hit),
        ("MAE", bt.mae),
        ("RMSE", bt.rmse),
    ]:
        ws5.append([label, val])
    ws5.append([])
    ws5.append(["Note: this backtests the model's own CAPM-signal methodology "
                "against realized outcomes - it is not a trading-strategy P&L backtest."])
    _autofit(ws5, 2)

    # --- Full fundamentals sheets (raw, as fetched - everything the
    # curated "Key Ratios" sheet above didn't have a named metric for) ---
    _add_raw_table_sheet(wb, "All Ratios (Raw)", fundamentals.get("ratios_5yr", {}))
    _add_raw_table_sheet(wb, "Profit & Loss", fundamentals.get("profit_loss", {}))
    _add_raw_table_sheet(wb, "Balance Sheet", fundamentals.get("balance_sheet", {}))
    _add_raw_table_sheet(wb, "Cash Flow", fundamentals.get("cash_flow", {}))

    # --- Raw Price Data sheet ---
    ws6 = wb.create_sheet("Raw Price Data (5yr)")
    ws6.append(["Date", "Open", "High", "Low", "Close", "Adj Close", "Volume"])
    _style_header_row(ws6, 1, 7)
    for date, row in price_df.iterrows():
        ws6.append([
            str(date.date()), row.get("Open"), row.get("High"), row.get("Low"),
            row.get("Close"), row.get("Adj Close"), row.get("Volume"),
        ])
    _autofit(ws6, 7)

    wb.save(output_path)
    return output_path
